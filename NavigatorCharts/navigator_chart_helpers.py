"""
Shared constants, matrix math, data loading, and Excel-writing utilities
for the navigator charts scripts and Streamlit app.
"""
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CALIBRATION_PARQUET = Path(__file__).parent / "navigator_chart_calibration_runs.parquet"

SPEEDS = [0, 15, 20, 25, 30, 35, 40, 45, 50]
STOP_SIGN_SECONDS = 15.0


# ── Matrix computation ────────────────────────────────────────────────────────

BLANK = None   # sentinel for cells that should be left empty


def _build_matrix(func):
    return [[func(in_s, out_s) for out_s in SPEEDS] for in_s in SPEEDS]


def matrix_transition(accel, decel):
    """Extra seconds for any speed change; blank on diagonal."""
    def f(in_s, out_s):
        if in_s == out_s:
            return BLANK
        if out_s > in_s:
            return accel[out_s] - accel[in_s]
        return decel[in_s] - decel[out_s]
    return _build_matrix(f)


def matrix_stop_go(accel, decel):
    """Standstill seconds remaining inside STOP_SIGN_SECONDS mandatory stop."""
    def f(in_s, out_s):
        if in_s == 0 and out_s == 0:
            return BLANK
        if in_s == 0:
            return accel[out_s]
        if out_s == 0:
            return decel[in_s]
        return STOP_SIGN_SECONDS - decel[in_s] - accel[out_s]
    return _build_matrix(f)


def matrix_turn_loss(accel, decel, ref_mph):
    """Extra seconds vs arriving and leaving at ref_mph; blank when entrance or exit < ref_mph."""
    def f(in_s, out_s):
        if in_s < ref_mph or out_s < ref_mph:
            return BLANK
        return (decel[in_s] - decel[ref_mph]) + (accel[out_s] - accel[ref_mph])
    return _build_matrix(f)


# ── Data loading / loss computation ──────────────────────────────────────────

def load_calibration_runs():
    """Return full normalized DataFrame from the calibration runs parquet."""
    df = pd.read_parquet(CALIBRATION_PARQUET)
    df["date"] = df["date"].astype(str).str[:10]
    return df


def compute_losses(df):
    """
    Return per-speed loss table (DataFrame) from all rows in df.
    Returns None if the three required test types are not all present.

    Columns: MPH, Straight (s), Accel Loss (s), Decel Loss (s)
    """
    if df.empty:
        return None
    # Group by course (notes) so losses are computed within matched courses only
    grp = (df.groupby(["notes", "test_type", "target_mph"])["time_s"]
             .mean()
             .unstack("test_type"))
    need = {"straight_speed", "start_speed", "speed_stop"}
    if not need.issubset(grp.columns):
        return None
    grp = grp.dropna()
    if grp.empty:
        return None
    per_course = pd.DataFrame({
        "Straight (s)":   grp["straight_speed"],
        "Accel Loss (s)": grp["start_speed"] - grp["straight_speed"],
        "Decel Loss (s)": grp["speed_stop"]  - grp["straight_speed"],
    })
    result = per_course.groupby("target_mph").mean().reset_index()
    return result.rename(columns={"target_mph": "MPH"})[
        ["MPH", "Straight (s)", "Accel Loss (s)", "Decel Loss (s)"]
    ]


def losses_to_dicts(losses):
    """
    Convert losses DataFrame to (accel, decel) dicts keyed by MPH.
    Missing speeds are filled with NaN.  Returns (None, None) if losses is None.
    """
    if losses is None:
        return None, None
    accel = {0: 0.0}
    decel = {0: 0.0}
    for mph, a, d in zip(losses["MPH"], losses["Accel Loss (s)"], losses["Decel Loss (s)"]):
        accel[int(mph)] = float(a)
        decel[int(mph)] = float(d)
    for s in SPEEDS:
        accel.setdefault(s, float("nan"))
        decel.setdefault(s, float("nan"))
    return accel, decel


# ── Excel helpers ─────────────────────────────────────────────────────────────

C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_AXIS_BG   = "2E75B6"
C_AXIS_FG   = "FFFFFF"
C_TITLE_BG  = "D6E4F0"
C_BORDER    = "9DC3E6"

BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
BLACK_FONT = Font(color="000000", size=10)


def header_style(bold=True):
    return {
        "fill":      PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid"),
        "font":      Font(color=C_HEADER_FG, bold=bold, size=10),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def axis_style():
    return {
        "fill":      PatternFill(start_color=C_AXIS_BG, end_color=C_AXIS_BG, fill_type="solid"),
        "font":      Font(color=C_AXIS_FG, bold=True, size=10),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def apply(cell, style_dict):
    for attr, val in style_dict.items():
        setattr(cell, attr, val)


def write_matrix(ws, matrix, title, subtitle, top_row, left_col,
                 color_lo, color_mid, color_hi, mid_value=0.0,
                 hide_zero_axis=False, color_scale=True):
    """
    Write one labeled matrix block to the worksheet.

    hide_zero_axis: paint the In=0 row and Out=0 column black-on-black so the
    values are hidden but still present.  The color scale is restricted to the
    remaining (In>0, Out>0) cells so it does not override the black fill.
    """
    n = len(SPEEDS)
    R, C = top_row, left_col

    # Title row
    title_cell = ws.cell(R, C, title)
    title_cell.font = Font(bold=True, size=11, color="1F4E79")
    title_cell.fill = PatternFill(start_color=C_TITLE_BG, end_color=C_TITLE_BG, fill_type="solid")
    ws.merge_cells(start_row=R, start_column=C, end_row=R, end_column=C + n)
    R += 1

    # Subtitle row
    sub_cell = ws.cell(R, C, subtitle)
    sub_cell.font = Font(italic=True, size=9, color="595959")
    ws.merge_cells(start_row=R, start_column=C, end_row=R, end_column=C + n)
    R += 1

    # Column headers
    corner = ws.cell(R, C, "In ↓  Out →")
    corner.fill = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid")
    corner.font = Font(color=C_HEADER_FG, bold=True, size=9)
    corner.alignment = Alignment(horizontal="center", vertical="center")

    for j, spd in enumerate(SPEEDS):
        cell = ws.cell(R, C + 1 + j, spd)
        if hide_zero_axis and j == 0:
            cell.fill = BLACK_FILL
            cell.font = BLACK_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            apply(cell, header_style())
    R += 1

    # Data rows
    data_start_row = R
    for i, in_spd in enumerate(SPEEDS):
        zero_row = hide_zero_axis and i == 0

        lbl = ws.cell(R + i, C, in_spd)
        if zero_row:
            lbl.fill = BLACK_FILL
            lbl.font = BLACK_FONT
            lbl.alignment = Alignment(horizontal="center", vertical="center")
        else:
            apply(lbl, axis_style())

        for j, val in enumerate(matrix[i]):
            cell = ws.cell(R + i, C + 1 + j)
            zero_col = hide_zero_axis and j == 0

            if zero_row or zero_col:
                cell.value = None if val is BLANK else round(val, 2)
                cell.fill = BLACK_FILL
                cell.font = BLACK_FONT
                if val is not BLANK:
                    cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center")
            elif val is BLANK:
                cell.value = None
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            else:
                cell.value = round(val, 2)
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()

    data_end_row = R + n - 1

    # Conditional color scale (excludes hidden axis when hide_zero_axis)
    cs_start_col = get_column_letter(C + (2 if hide_zero_axis else 1))
    cs_end_col   = get_column_letter(C + n)
    cs_start_row = data_start_row + (1 if hide_zero_axis else 0)
    cs_range = f"{cs_start_col}{cs_start_row}:{cs_end_col}{data_end_row}"

    if color_scale:
        rule = ColorScaleRule(
            start_type="min",  start_color=color_lo,
            mid_type="num",    mid_value=mid_value, mid_color=color_mid,
            end_type="max",    end_color=color_hi,
        )
        ws.conditional_formatting.add(cs_range, rule)

    return R + n  # next available row


def write_reference_charts_to_sheet(ws, accel, decel, label="", color_scale=True):
    """
    Write 4 reference matrices directly into an existing openpyxl worksheet.
    label: optional string appended to each matrix title.
    """
    ws.column_dimensions["A"].width = 10
    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 8
    for r in range(1, 80):
        ws.row_dimensions[r].height = 16

    tag = f"  [{label}]" if label else ""

    m1 = matrix_transition(accel, decel)
    m2 = matrix_stop_go(accel, decel)
    m3 = matrix_turn_loss(accel, decel, ref_mph=15)
    m4 = matrix_turn_loss(accel, decel, ref_mph=20)

    next_row = write_matrix(
        ws, m1,
        title=f"1.  Speed Transition Time (seconds){tag}",
        subtitle="Extra seconds vs. flying straight through · blank diagonal = no change",
        top_row=1, left_col=1,
        color_lo="63BE7B", color_mid="FFEB84", color_hi="F8696B",
        mid_value=0.0,
        color_scale=color_scale,
    )
    next_row = write_matrix(
        ws, m2,
        title=f"2.  Available Pause at Stop Sign (seconds){tag}",
        subtitle=(
            f"Standstill seconds remaining inside a {int(STOP_SIGN_SECONDS)}-second "
            "mandatory stop · Out=0 col = raw brake cost"
        ),
        top_row=next_row + 2, left_col=1,
        color_lo="63BE7B", color_mid="FFEB84", color_hi="F8696B",
        mid_value=STOP_SIGN_SECONDS / 2,
        hide_zero_axis=True,
        color_scale=color_scale,
    )
    next_row = write_matrix(
        ws, m3,
        title=f"3.  Turn Time Lost vs. 15 mph reference (seconds){tag}",
        subtitle="Extra seconds vs. In=15→15 · cells where In or Out < 15 mph are blank",
        top_row=next_row + 2, left_col=1,
        color_lo="63BE7B", color_mid="FFEB84", color_hi="F8696B",
        mid_value=0.0,
        hide_zero_axis=True,
        color_scale=color_scale,
    )
    next_row = write_matrix(
        ws, m4,
        title=f"4.  Turn Time Lost vs. 20 mph reference (seconds){tag}",
        subtitle="Extra seconds vs. In=20→20 · cells where In or Out < 20 mph are blank",
        top_row=next_row + 2, left_col=1,
        color_lo="63BE7B", color_mid="FFEB84", color_hi="F8696B",
        mid_value=0.0,
        hide_zero_axis=True,
        color_scale=color_scale,
    )

    note_row = next_row + 2
    note = ws.cell(note_row, 1,
                   "Source: latest calibration runs" + (f" · {label}" if label else ""))
    note.font = Font(italic=True, size=8, color="808080")


def build_reference_workbook(accel, decel, label="", color_scale=True):
    """Return a new openpyxl Workbook with 4 reference matrices (for in-memory export)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Navigator Charts"
    write_reference_charts_to_sheet(ws, accel, decel, label, color_scale=color_scale)
    return wb


def _poly_eq_str(coef, var="x"):
    a2, a1, a0 = coef
    parts = [
        f"{a2:.6f}{var}²",
        f"{'+' if a1 >= 0 else '-'} {abs(a1):.6f}{var}",
        f"{'+' if a0 >= 0 else '-'} {abs(a0):.6f}",
    ]
    return " ".join(parts)


def _r2_poly(y, coef, x):
    ss_res = float(np.sum((y - np.polyval(coef, x)) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    return 1.0 - ss_res / ss_tot if ss_tot > 0 else 1.0


def _write_reference_section(ws, start_row, losses, ca, cd, df_runs, label):
    """Write losses table, polynomial fit equations, and raw run data below the matrices."""
    r = start_row

    title_cell = ws.cell(r, 1, f"Reference Data — {label}")
    title_cell.font = Font(bold=True, size=11, color="1F4E79")
    r += 2

    if losses is not None and not losses.empty:
        has_fit = ca is not None and cd is not None
        hdrs = ["MPH", "Accel Loss (s)", "Decel Loss (s)"]
        if has_fit:
            hdrs += ["Fit Accel (s)", "Fit Decel (s)"]
        for j, h in enumerate(hdrs):
            c = ws.cell(r, 1 + j, h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid")
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()
        r += 1

        x = losses["MPH"].values.astype(float)
        fit_a = np.polyval(ca, x) if has_fit else None
        fit_d = np.polyval(cd, x) if has_fit else None

        for i, (_, row_data) in enumerate(losses.iterrows()):
            ws.cell(r, 1, int(row_data["MPH"]))
            ws.cell(r, 2, round(float(row_data["Accel Loss (s)"]), 3))
            ws.cell(r, 3, round(float(row_data["Decel Loss (s)"]), 3))
            if has_fit:
                ws.cell(r, 4, round(float(fit_a[i]), 3))
                ws.cell(r, 5, round(float(fit_d[i]), 3))
            r += 1
        r += 1

        if has_fit:
            r2a = _r2_poly(losses["Accel Loss (s)"].values, ca, x)
            r2d = _r2_poly(losses["Decel Loss (s)"].values, cd, x)
            eq_a = ws.cell(r, 1, f"Accel fit (R²={r2a:.4f}): {_poly_eq_str(ca)}")
            eq_a.font = Font(italic=True, size=9, color="595959")
            r += 1
            eq_d = ws.cell(r, 1, f"Decel fit (R²={r2d:.4f}): {_poly_eq_str(cd)}")
            eq_d.font = Font(italic=True, size=9, color="595959")
            r += 2

    if df_runs is not None and not df_runs.empty:
        cell = ws.cell(r, 1, "Calibration Runs")
        cell.font = Font(bold=True, size=10, color="1F4E79")
        r += 1

        run_cols = [c for c in
                    ["excluded", "date", "notes", "target_mph", "test_type", "run_number", "time_s", "direction"]
                    if c in df_runs.columns]
        for j, h in enumerate(run_cols):
            c = ws.cell(r, 1 + j, h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill(start_color=C_AXIS_BG, end_color=C_AXIS_BG, fill_type="solid")
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border()
        r += 1

        _EXCL_FILL = PatternFill(start_color="FFFFD9D9", end_color="FFFFD9D9", fill_type="solid")
        _has_excl = "excluded" in run_cols

        for _, row_data in df_runs.iterrows():
            is_excluded = _has_excl and bool(row_data["excluded"])
            for j, col in enumerate(run_cols):
                val = row_data[col]
                if hasattr(val, "item"):
                    val = val.item()
                cell = ws.cell(r, 1 + j, val)
                if is_excluded:
                    cell.fill = _EXCL_FILL
            r += 1


def build_combined_workbook(
    accel_all, decel_all, losses_all, ca_all, cd_all, df_all_runs,
    accel_filt, decel_filt, losses_filt, ca_filt, cd_filt, df_filt_runs,
    color_scale=True,
):
    """Return a two-sheet workbook: All Data and Filtered, each with matrices + reference data."""
    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "All Data"
    if accel_all is not None:
        write_reference_charts_to_sheet(ws_all, accel_all, decel_all, "all data",
                                        color_scale=color_scale)
        _write_reference_section(ws_all, 60, losses_all, ca_all, cd_all, df_all_runs, "All Data")

    ws_filt = wb.create_sheet("Filtered")
    if accel_filt is not None:
        write_reference_charts_to_sheet(ws_filt, accel_filt, decel_filt, "filtered",
                                        color_scale=color_scale)
        _write_reference_section(ws_filt, 60, losses_filt, ca_filt, cd_filt, df_filt_runs,
                                 "Filtered")

    return wb
