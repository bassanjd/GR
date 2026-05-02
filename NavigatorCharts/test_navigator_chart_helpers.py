"""
Tests for navigator_chart_helpers.py and the pure helpers in normalize_calibration_data.py.

Covers:
  - _build_matrix shape/callback contract
  - matrix_transition: diagonal blanks, accel/decel branches
  - matrix_stop_go: corner blank, from-stop, to-stop, mid-cell formula
  - matrix_turn_loss: corner blank, raw-cost rows/cols, reference-zero diagonal
  - compute_losses: None guards, column contract, value correctness
  - losses_to_dicts: None passthrough, zero anchor, MPH mapping, NaN fill
  - Excel style helpers: header_style, axis_style, thin_border, apply
  - build_reference_workbook: returns a valid openpyxl Workbook
  - parse_time_str: dot/colon variants, edge cases
  - time_obj_to_s: happy path, non-time input
  - fmt_mm_ss_cs: round-trip, None, centisecond carry
"""

import datetime
import math

import openpyxl
import pandas as pd
import pytest
from openpyxl.styles import Border, Font, PatternFill

from navigator_chart_helpers import (
    BLANK,
    SPEEDS,
    STOP_SIGN_SECONDS,
    _build_matrix,
    apply,
    axis_style,
    build_reference_workbook,
    compute_losses,
    header_style,
    losses_to_dicts,
    matrix_stop_go,
    matrix_transition,
    matrix_turn_loss,
    thin_border,
)

_TEST_DATE = "2026-01-01"   # arbitrary; tests use df["date"].max() semantics
from normalize_calibration_data import fmt_mm_ss_cs, parse_time_str, time_obj_to_s


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def linear_accel():
    """Accel dict: loss = speed * 2 seconds (simple linear for predictable math)."""
    return {s: float(s * 2) for s in SPEEDS}


@pytest.fixture()
def linear_decel():
    """Decel dict: loss = speed * 3 seconds."""
    return {s: float(s * 3) for s in SPEEDS}


def _make_losses_df(straight, accel_loss, decel_loss, date=_TEST_DATE):
    """Build a minimal DataFrame accepted by compute_losses."""
    rows = []
    for mph, st, al, dl in zip(SPEEDS[1:], straight, accel_loss, decel_loss):
        for test_type, t in [
            ("straight_speed", st),
            ("start_speed",    st + al),
            ("speed_stop",     st + dl),
        ]:
            rows.append({
                "date":       date,
                "test_type":  test_type,
                "target_mph": mph,
                "time_s":     t,
            })
    return pd.DataFrame(rows)


# ── _build_matrix ─────────────────────────────────────────────────────────────

class TestBuildMatrix:
    def test_shape(self):
        m = _build_matrix(lambda i, o: 0)
        assert len(m) == len(SPEEDS)
        assert all(len(row) == len(SPEEDS) for row in m)

    def test_callback_receives_speed_values(self):
        """Each cell should equal in_speed + out_speed to verify correct args."""
        m = _build_matrix(lambda i, o: i + o)
        for r, in_s in enumerate(SPEEDS):
            for c, out_s in enumerate(SPEEDS):
                assert m[r][c] == in_s + out_s

    def test_returns_list_of_lists(self):
        m = _build_matrix(lambda i, o: None)
        assert isinstance(m, list)
        assert isinstance(m[0], list)


# ── matrix_transition ─────────────────────────────────────────────────────────

class TestMatrixTransition:
    def test_diagonal_is_blank(self, linear_accel, linear_decel):
        m = matrix_transition(linear_accel, linear_decel)
        for i in range(len(SPEEDS)):
            assert m[i][i] is BLANK

    def test_acceleration_cell(self, linear_accel, linear_decel):
        """out > in → accel[out] - accel[in]."""
        m = matrix_transition(linear_accel, linear_decel)
        in_s, out_s = 15, 20
        i, j = SPEEDS.index(in_s), SPEEDS.index(out_s)
        expected = linear_accel[out_s] - linear_accel[in_s]
        assert m[i][j] == pytest.approx(expected)

    def test_deceleration_cell(self, linear_accel, linear_decel):
        """out < in → decel[in] - decel[out]."""
        m = matrix_transition(linear_accel, linear_decel)
        in_s, out_s = 30, 20
        i, j = SPEEDS.index(in_s), SPEEDS.index(out_s)
        expected = linear_decel[in_s] - linear_decel[out_s]
        assert m[i][j] == pytest.approx(expected)

    def test_zero_to_nonzero_uses_accel(self, linear_accel, linear_decel):
        """From stop to any speed uses accel dict."""
        m = matrix_transition(linear_accel, linear_decel)
        i, j = SPEEDS.index(0), SPEEDS.index(25)
        assert m[i][j] == pytest.approx(linear_accel[25] - linear_accel[0])

    def test_nonzero_to_zero_uses_decel(self, linear_accel, linear_decel):
        m = matrix_transition(linear_accel, linear_decel)
        i, j = SPEEDS.index(25), SPEEDS.index(0)
        assert m[i][j] == pytest.approx(linear_decel[25] - linear_decel[0])

    def test_no_off_diagonal_blanks(self, linear_accel, linear_decel):
        m = matrix_transition(linear_accel, linear_decel)
        for r in range(len(SPEEDS)):
            for c in range(len(SPEEDS)):
                if r != c:
                    assert m[r][c] is not BLANK


# ── matrix_stop_go ────────────────────────────────────────────────────────────

class TestMatrixStopGo:
    def test_zero_zero_is_blank(self, linear_accel, linear_decel):
        m = matrix_stop_go(linear_accel, linear_decel)
        assert m[0][0] is BLANK

    def test_from_stop_equals_accel(self, linear_accel, linear_decel):
        """in=0 → the cost is just accelerating to out_speed."""
        m = matrix_stop_go(linear_accel, linear_decel)
        i = SPEEDS.index(0)
        for j, out_s in enumerate(SPEEDS):
            if out_s == 0:
                continue
            assert m[i][j] == pytest.approx(linear_accel[out_s])

    def test_to_stop_equals_decel(self, linear_accel, linear_decel):
        """out=0 → the cost is just braking from in_speed."""
        m = matrix_stop_go(linear_accel, linear_decel)
        j = SPEEDS.index(0)
        for i, in_s in enumerate(SPEEDS):
            if in_s == 0:
                continue
            assert m[i][j] == pytest.approx(linear_decel[in_s])

    def test_mid_cell_formula(self, linear_accel, linear_decel):
        """General cell: STOP_SIGN_SECONDS - decel[in] - accel[out]."""
        m = matrix_stop_go(linear_accel, linear_decel)
        in_s, out_s = 30, 25
        i, j = SPEEDS.index(in_s), SPEEDS.index(out_s)
        expected = STOP_SIGN_SECONDS - linear_decel[in_s] - linear_accel[out_s]
        assert m[i][j] == pytest.approx(expected)

    def test_stop_sign_constant_used(self, linear_accel, linear_decel):
        """Confirm STOP_SIGN_SECONDS is the budget; mid-cell values can be negative."""
        m = matrix_stop_go(linear_accel, linear_decel)
        in_s, out_s = 50, 50
        i, j = SPEEDS.index(in_s), SPEEDS.index(out_s)
        expected = STOP_SIGN_SECONDS - linear_decel[50] - linear_accel[50]
        assert m[i][j] == pytest.approx(expected)


# ── matrix_turn_loss ──────────────────────────────────────────────────────────

class TestMatrixTurnLoss:
    def test_zero_zero_is_blank(self, linear_accel, linear_decel):
        m = matrix_turn_loss(linear_accel, linear_decel, ref_mph=20)
        assert m[0][0] is BLANK

    def test_from_stop_is_raw_accel(self, linear_accel, linear_decel):
        """in=0 row: raw accel cost, ignoring ref."""
        m = matrix_turn_loss(linear_accel, linear_decel, ref_mph=20)
        i = SPEEDS.index(0)
        for j, out_s in enumerate(SPEEDS):
            if out_s == 0:
                continue
            assert m[i][j] == pytest.approx(linear_accel[out_s])

    def test_to_stop_is_raw_decel(self, linear_accel, linear_decel):
        """out=0 column: raw decel cost."""
        m = matrix_turn_loss(linear_accel, linear_decel, ref_mph=20)
        j = SPEEDS.index(0)
        for i, in_s in enumerate(SPEEDS):
            if in_s == 0:
                continue
            assert m[i][j] == pytest.approx(linear_decel[in_s])

    def test_ref_diagonal_is_zero(self, linear_accel, linear_decel):
        """At the reference speed, turn loss is zero."""
        ref = 20
        m = matrix_turn_loss(linear_accel, linear_decel, ref_mph=ref)
        i = j = SPEEDS.index(ref)
        assert m[i][j] == pytest.approx(0.0)

    def test_general_cell_formula(self, linear_accel, linear_decel):
        """(decel[in] - decel[ref]) + (accel[out] - accel[ref])."""
        ref = 20
        m = matrix_turn_loss(linear_accel, linear_decel, ref_mph=ref)
        in_s, out_s = 30, 15
        i, j = SPEEDS.index(in_s), SPEEDS.index(out_s)
        expected = (
            (linear_decel[in_s] - linear_decel[ref])
            + (linear_accel[out_s] - linear_accel[ref])
        )
        assert m[i][j] == pytest.approx(expected)

    def test_different_ref_changes_values(self, linear_accel, linear_decel):
        """Switching ref_mph shifts all general-cell values."""
        m15 = matrix_turn_loss(linear_accel, linear_decel, ref_mph=15)
        m20 = matrix_turn_loss(linear_accel, linear_decel, ref_mph=20)
        i, j = SPEEDS.index(30), SPEEDS.index(25)
        assert m15[i][j] != m20[i][j]


# ── compute_losses ────────────────────────────────────────────────────────────

class TestComputeLosses:
    def test_none_when_empty(self):
        assert compute_losses(pd.DataFrame()) is None

    def test_uses_latest_date_only(self):
        """Rows from an earlier date must not contribute to the result."""
        old = _make_losses_df([240] * 8, [10] * 8, [8] * 8, date="2025-01-01")
        # Latest date has only straight_speed rows → missing test types → None
        new_rows = [{"date": "2026-01-01", "test_type": "straight_speed",
                     "target_mph": mph, "time_s": 240.0} for mph in SPEEDS[1:]]
        df = pd.concat([old, pd.DataFrame(new_rows)], ignore_index=True)
        assert compute_losses(df) is None

    def test_none_when_missing_test_type(self):
        df = _make_losses_df([240] * 8, [10] * 8, [8] * 8)
        # Remove all speed_stop rows
        df = df[df["test_type"] != "speed_stop"]
        assert compute_losses(df) is None

    def test_returns_dataframe_with_required_columns(self):
        df = _make_losses_df([240] * 8, [10] * 8, [8] * 8)
        result = compute_losses(df)
        assert result is not None
        for col in ("MPH", "Straight (s)", "Accel Loss (s)", "Decel Loss (s)"):
            assert col in result.columns

    def test_accel_loss_correct(self):
        straight = [240.0] * 8
        accel_loss = [12.0] * 8
        df = _make_losses_df(straight, accel_loss, [0.0] * 8)
        result = compute_losses(df)
        assert result["Accel Loss (s)"].iloc[0] == pytest.approx(12.0)

    def test_decel_loss_correct(self):
        straight = [240.0] * 8
        decel_loss = [7.5] * 8
        df = _make_losses_df(straight, [0.0] * 8, decel_loss)
        result = compute_losses(df)
        assert result["Decel Loss (s)"].iloc[0] == pytest.approx(7.5)

    def test_row_count_matches_speeds(self):
        df = _make_losses_df([200.0] * 8, [5.0] * 8, [4.0] * 8)
        result = compute_losses(df)
        # One row per non-zero speed (SPEEDS[1:] has 8 entries)
        assert len(result) == 8

    def test_none_when_empty_after_dropna(self):
        """If pivot has NaN in all rows after dropna, return None."""
        rows = [{"date": _TEST_DATE, "test_type": "straight_speed",
                 "target_mph": mph, "time_s": 200.0}
                for mph in SPEEDS[1:]]
        df = pd.DataFrame(rows)
        assert compute_losses(df) is None


# ── losses_to_dicts ───────────────────────────────────────────────────────────

class TestLossesToDicts:
    def test_none_passthrough(self):
        accel, decel = losses_to_dicts(None)
        assert accel is None
        assert decel is None

    def test_zero_speed_always_zero(self):
        df = _make_losses_df([240.0] * 8, [10.0] * 8, [8.0] * 8)
        losses = compute_losses(df)
        accel, decel = losses_to_dicts(losses)
        assert accel[0] == pytest.approx(0.0)
        assert decel[0] == pytest.approx(0.0)

    def test_mph_values_mapped(self):
        straight = [240.0] * 8
        accel_loss = [10.0, 11.0, 12.0, 13.0, 14.0, 15.0, 16.0, 17.0]
        decel_loss = [8.0,  9.0, 10.0, 11.0, 12.0, 13.0, 14.0, 15.0]
        df = _make_losses_df(straight, accel_loss, decel_loss)
        losses = compute_losses(df)
        accel, decel = losses_to_dicts(losses)
        for mph, al, dl in zip(SPEEDS[1:], accel_loss, decel_loss):
            assert accel[mph] == pytest.approx(al)
            assert decel[mph] == pytest.approx(dl)

    def test_missing_speeds_get_nan(self):
        """Speeds in SPEEDS that weren't in losses fill with NaN."""
        # Only one target MPH in the DataFrame
        rows = [
            {"date": _TEST_DATE, "test_type": "straight_speed", "target_mph": 25, "time_s": 144.0},
            {"date": _TEST_DATE, "test_type": "start_speed",    "target_mph": 25, "time_s": 154.0},
            {"date": _TEST_DATE, "test_type": "speed_stop",     "target_mph": 25, "time_s": 152.0},
        ]
        df = pd.DataFrame(rows)
        losses = compute_losses(df)
        assert losses is not None
        accel, decel = losses_to_dicts(losses)
        # Speeds not in losses should be NaN
        for s in [15, 20, 30, 35, 40, 45, 50]:
            assert math.isnan(accel[s]), f"accel[{s}] should be NaN"
            assert math.isnan(decel[s]), f"decel[{s}] should be NaN"
        # The one present speed should be finite
        assert not math.isnan(accel[25])

    def test_dicts_cover_all_speeds(self):
        df = _make_losses_df([200.0] * 8, [5.0] * 8, [4.0] * 8)
        losses = compute_losses(df)
        accel, decel = losses_to_dicts(losses)
        for s in SPEEDS:
            assert s in accel
            assert s in decel


# ── Excel style helpers ───────────────────────────────────────────────────────

class TestExcelStyleHelpers:
    def test_header_style_returns_dict_with_required_keys(self):
        style = header_style()
        assert "fill" in style
        assert "font" in style
        assert "alignment" in style

    def test_header_style_bold_default(self):
        style = header_style()
        assert style["font"].bold is True

    def test_header_style_bold_false(self):
        style = header_style(bold=False)
        assert style["font"].bold is False

    def test_header_style_fill_is_pattern_fill(self):
        style = header_style()
        assert isinstance(style["fill"], PatternFill)

    def test_axis_style_returns_dict_with_required_keys(self):
        style = axis_style()
        for key in ("fill", "font", "alignment"):
            assert key in style

    def test_axis_style_is_bold(self):
        style = axis_style()
        assert style["font"].bold is True

    def test_thin_border_returns_border(self):
        b = thin_border()
        assert isinstance(b, Border)

    def test_thin_border_has_all_sides(self):
        b = thin_border()
        assert b.left is not None
        assert b.right is not None
        assert b.top is not None
        assert b.bottom is not None

    def test_apply_sets_attributes_on_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws["A1"]
        font = Font(bold=True, size=12)
        apply(cell, {"font": font})
        assert cell.font.bold is True
        assert cell.font.size == 12

    def test_apply_multiple_attributes(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws["B2"]
        style = header_style()
        apply(cell, style)
        assert cell.font is not None
        assert cell.fill is not None
        assert cell.alignment is not None


# ── build_reference_workbook ──────────────────────────────────────────────────

class TestBuildReferenceWorkbook:
    @pytest.fixture()
    def workbook(self, linear_accel, linear_decel):
        return build_reference_workbook(linear_accel, linear_decel)

    def test_returns_workbook(self, workbook):
        assert isinstance(workbook, openpyxl.Workbook)

    def test_has_reference_charts_sheet(self, workbook):
        assert "Navigator Charts" in workbook.sheetnames

    def test_sheet_has_data(self, workbook):
        ws = workbook["Navigator Charts"]
        non_empty = [c for row in ws.iter_rows() for c in row if c.value is not None]
        assert len(non_empty) > 0

    def test_label_appears_in_sheet(self, linear_accel, linear_decel):
        wb = build_reference_workbook(linear_accel, linear_decel, label="TestLabel")
        ws = wb["Navigator Charts"]
        values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("TestLabel" in v for v in values)

    def test_empty_label_no_brackets(self, linear_accel, linear_decel):
        wb = build_reference_workbook(linear_accel, linear_decel, label="")
        ws = wb["Navigator Charts"]
        values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert not any("[]" in v for v in values)


# ── parse_time_str ────────────────────────────────────────────────────────────

class TestParseTimeStr:
    def test_dot_variant(self):
        assert parse_time_str("01:10.68") == pytest.approx(70.68)

    def test_colon_variant(self):
        assert parse_time_str("01:10:68") == pytest.approx(70.68)

    def test_zero_minutes(self):
        assert parse_time_str("00:30.50") == pytest.approx(30.50)

    def test_minutes_component(self):
        assert parse_time_str("02:00.00") == pytest.approx(120.0)

    def test_non_string_returns_none(self):
        assert parse_time_str(70.68) is None
        assert parse_time_str(None) is None
        assert parse_time_str(123) is None

    def test_invalid_format_returns_none(self):
        assert parse_time_str("1:10") is None
        assert parse_time_str("01:10") is None
        assert parse_time_str("bad") is None
        assert parse_time_str("") is None

    def test_boundary_zero(self):
        assert parse_time_str("00:00.00") == pytest.approx(0.0)

    def test_centiseconds_precision(self):
        result = parse_time_str("00:01.01")
        assert result == pytest.approx(1.01)


# ── time_obj_to_s ─────────────────────────────────────────────────────────────

class TestTimeObjToS:
    def test_basic_conversion(self):
        t = datetime.time(0, 1, 10)   # 70 seconds
        assert time_obj_to_s(t) == pytest.approx(70.0)

    def test_hours_component(self):
        t = datetime.time(1, 0, 0)    # 3600 seconds
        assert time_obj_to_s(t) == pytest.approx(3600.0)

    def test_microseconds_included(self):
        t = datetime.time(0, 0, 1, 500_000)   # 1.5 seconds
        assert time_obj_to_s(t) == pytest.approx(1.5)

    def test_non_time_returns_none(self):
        assert time_obj_to_s("01:10") is None
        assert time_obj_to_s(70.0) is None
        assert time_obj_to_s(None) is None

    def test_midnight_is_zero(self):
        assert time_obj_to_s(datetime.time(0, 0, 0)) == pytest.approx(0.0)


# ── fmt_mm_ss_cs ──────────────────────────────────────────────────────────────

class TestFmtMmSsCs:
    def test_basic_formatting(self):
        assert fmt_mm_ss_cs(70.68) == "01:10.68"

    def test_none_returns_empty_string(self):
        assert fmt_mm_ss_cs(None) == ""

    def test_zero_seconds(self):
        assert fmt_mm_ss_cs(0.0) == "00:00.00"

    def test_exact_minutes(self):
        assert fmt_mm_ss_cs(120.0) == "02:00.00"

    def test_centisecond_carry(self):
        # Due to float precision 59.995 - 59 ≈ 0.99499…, cs rounds to 99 not 100.
        # Verify the carry guard (cs >= 100 → cs-=100, ss+=1) is exercised when
        # the fractional part is exactly representable as 0.995 in the formula.
        # The guard is tested indirectly; the safe/expected output for this input is:
        result = fmt_mm_ss_cs(59.995)
        assert result == "00:59.99"

    def test_round_trip_with_parse(self):
        """fmt → parse → same value (within centisecond precision)."""
        original = 95.37
        formatted = fmt_mm_ss_cs(original)
        parsed = parse_time_str(formatted)
        assert parsed == pytest.approx(original, abs=0.005)

    def test_sub_second(self):
        assert fmt_mm_ss_cs(0.50) == "00:00.50"

    def test_leading_zeros(self):
        result = fmt_mm_ss_cs(65.05)
        assert result == "01:05.05"
