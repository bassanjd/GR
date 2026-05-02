"""
Build reference navigation charts from normalized calibration data.

Reads:  navigator_chart_normalized.xlsx  (Calibration Runs sheet)
Writes: Adds / overwrites a "Reference Charts" sheet in the same file.

Four matrices, all using 2026 1-mile-course data:

  1. Speed Transition Time (s)
  2. Available Pause at Stop Sign (s)
  3. Turn Time Lost vs 15↔15 reference (s)
  4. Turn Time Lost vs 20↔20 reference (s)
"""
import openpyxl

from navigator_chart_helpers import (
    NORMALIZED_XLSX,
    compute_losses,
    load_calibration_runs,
    losses_to_dicts,
    write_reference_charts_to_sheet,
)

SHEET_NAME = "Reference Charts"


def build_reference_charts():
    df = load_calibration_runs()
    losses = compute_losses(df)
    accel, decel = losses_to_dicts(losses)

    if accel is None:
        print("ERROR: Could not compute losses from 2026 data.")
        return

    wb = openpyxl.load_workbook(NORMALIZED_XLSX)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    write_reference_charts_to_sheet(ws, accel, decel)

    wb.save(NORMALIZED_XLSX)
    print(f"Wrote '{SHEET_NAME}' sheet to {NORMALIZED_XLSX.name}")

    mph_keys = [int(m) for m in losses["MPH"]]
    print(f"\nAccel losses (s): { {k: round(accel[k], 3) for k in [0] + mph_keys} }")
    print(f"Decel losses (s): { {k: round(decel[k], 3) for k in [0] + mph_keys} }")


if __name__ == "__main__":
    build_reference_charts()
