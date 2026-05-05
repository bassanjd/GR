"""
Normalize navigator chart timing data to a parquet file.

Reads:  2026 Great Race Charts April 29th.xlsx
Writes: navigator_chart_calibration_runs.parquet

Three test types over a fixed measured course:
  straight_speed  — flying start and finish (baseline)
  start_speed     — standing start to flying finish
  speed_stop      — flying start to full stop

Time entry format in source data:
  "MM:SS:cs"  colon before centiseconds (common entry typo)
  "MM:SS.cs"  dot before centiseconds (intended format)
  Both encode: MM minutes, SS whole seconds, cs hundredths of a second.
  Example: "01:10:68" and "01:10.68" both mean 1 min 10.68 sec = 70.68 s.
"""
import datetime
import re
from pathlib import Path

import openpyxl
import pandas as pd

_MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def _date_from_filename(path):
    """Parse ISO date from a filename like '2026 Great Race Charts April 29th.xlsx'."""
    stem = Path(path).stem
    m = re.search(r'(\d{4}).*\b([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?$', stem)
    if not m:
        raise ValueError(f"Cannot parse calibration date from filename: {stem!r}")
    year, month_str, day = int(m.group(1)), m.group(2).lower()[:3], int(m.group(3))
    return f"{year}-{_MONTH_MAP[month_str]:02d}-{day:02d}"


def _col_dates(ws, col_start, col_end):
    """Read ISO date strings from row 2 for columns col_start..col_end (inclusive)."""
    return {
        col: ws.cell(2, col).value.strftime('%Y-%m-%d')
        for col in range(col_start, col_end + 1)
        if isinstance(ws.cell(2, col).value, (datetime.date, datetime.datetime))
    }


EXCEL_IN = Path(__file__).parent / "2026 Great Race Charts April 29th.xlsx"
PARQUET_OUT = Path(__file__).parent / "navigator_chart_calibration_runs.parquet"

SPEEDS_MPH = [15, 20, 25, 30, 35, 40, 45, 50]
DIRECTIONS_ALT = ['E', 'W', 'E', 'W', 'E', 'W', 'E', 'W']


# ── Parsing helpers ───────────────────────────────────────────────────────────

def parse_time_str(s):
    """Parse 'MM:SS:cs' or 'MM:SS.cs' string to decimal seconds.

    The colon variant (MM:SS:cs) is a common entry typo for the dot variant.
    Last two digits are always hundredths of a second in both formats.
    """
    if not isinstance(s, str):
        return None
    m = re.match(r'^(\d{1,2}):(\d{2})[:.](\d{2})$', s.strip())
    if m:
        mm, ss, cs = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return mm * 60 + ss + cs / 100
    return None


def time_obj_to_s(t):
    """Convert datetime.time to decimal seconds."""
    if not isinstance(t, datetime.time):
        return None
    return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1_000_000


def fmt_mm_ss_cs(total_s):
    """Format decimal seconds as 'MM:SS.cs' (normalized dot form)."""
    if total_s is None:
        return ""
    mm = int(total_s // 60)
    rem = total_s % 60
    ss = int(rem)
    cs = round((rem - ss) * 100)
    if cs >= 100:
        cs -= 100
        ss += 1
    return f"{mm:02d}:{ss:02d}.{cs:02d}"


# ── Data extraction ───────────────────────────────────────────────────────────

def extract_rows():
    """Return a list of dicts, one per timing run."""
    wb_f = openpyxl.load_workbook(EXCEL_IN, data_only=False)
    wb_v = openpyxl.load_workbook(EXCEL_IN, data_only=True)
    rows = []

    def append(date, test_type, mph, run_num, direction, t_s, raw, notes):
        if t_s is None:
            return
        rows.append({
            "date": date,
            "test_type": test_type,
            "target_mph": mph,
            "run_number": run_num,
            "direction": direction,
            "time_raw": raw,
            "time_s": round(t_s, 2),
            "notes": notes,
        })

    new_date = _date_from_filename(EXCEL_IN)

    # ── Straight Speed ────────────────────────────────────────────────────────

    # Old data (rows 3–10, datetime.time values, cols B–H); dates from row 2
    ws = wb_v["Straight Speed"]
    col_dates = _col_dates(ws, 2, 8)
    for row, mph in zip(range(3, 11), SPEEDS_MPH):
        for col in range(2, 9):          # B–H = runs 1–7
            val = ws.cell(row, col).value
            if not isinstance(val, datetime.time):
                continue
            t_s = time_obj_to_s(val)
            append(col_dates[col], "straight_speed", mph,
                   col - 1, "", t_s, fmt_mm_ss_cs(t_s), "short course")

    # New data (rows 21–28, raw "MM:SS:cs" strings, cols B–G = runs 1–6)
    ws_f = wb_f["Straight Speed"]
    for row, mph in zip(range(21, 29), SPEEDS_MPH):
        for col in range(2, 8):          # B–G = runs 1–6
            val = ws_f.cell(row, col).value
            t_s = parse_time_str(val)
            append(new_date, "straight_speed", mph,
                   col - 1, DIRECTIONS_ALT[col - 2], t_s,
                   str(val) if val else "", "")

    # ── Speed Stop ────────────────────────────────────────────────────────────

    # Old data (rows 3–10, datetime.time, cols B–H = runs 1–7); dates from row 2
    ws = wb_v["Speed Stop"]
    col_dates = _col_dates(ws, 2, 8)
    for row, mph in zip(range(3, 11), SPEEDS_MPH):
        for col in range(2, 9):
            val = ws.cell(row, col).value
            if not isinstance(val, datetime.time):
                continue
            t_s = time_obj_to_s(val)
            append(col_dates[col], "speed_stop", mph,
                   col - 1, "", t_s, fmt_mm_ss_cs(t_s), "short course")

    # New data (rows 33–40, raw strings, cols B–G = runs 1–6; H is AVG)
    ws_f = wb_f["Speed Stop"]
    for row, mph in zip(range(33, 41), SPEEDS_MPH):
        for col in range(2, 8):
            val = ws_f.cell(row, col).value
            t_s = parse_time_str(val)
            append(new_date, "speed_stop", mph,
                   col - 1, DIRECTIONS_ALT[col - 2], t_s,
                   str(val) if val else "", "")

    # ── Start Speed ───────────────────────────────────────────────────────────

    # Old data (rows 3–10, datetime.time, cols B–G = runs 1–6; H is AVG); dates from row 2
    ws = wb_v["Start Speed"]
    col_dates = _col_dates(ws, 2, 7)
    for row, mph in zip(range(3, 11), SPEEDS_MPH):
        for col in range(2, 8):
            val = ws.cell(row, col).value
            if not isinstance(val, datetime.time):
                continue
            t_s = time_obj_to_s(val)
            append(col_dates[col], "start_speed", mph,
                   col - 1, "", t_s, fmt_mm_ss_cs(t_s), "short course")

    # New data (rows 34–41, raw strings, cols B–G = runs 1–6; H is AVG)
    ws_f = wb_f["Start Speed"]
    for row, mph in zip(range(34, 42), SPEEDS_MPH):
        for col in range(2, 8):
            val = ws_f.cell(row, col).value
            t_s = parse_time_str(val)
            append(new_date, "start_speed", mph,
                   col - 1, DIRECTIONS_ALT[col - 2], t_s,
                   str(val) if val else "", "")

    return rows


# ── Derived calculations ──────────────────────────────────────────────────────

def compute_summary(df):
    """Per (date, test_type, target_mph): average time and std dev."""
    agg = (df.groupby(["date", "test_type", "target_mph"])["time_s"]
             .agg(avg_time_s="mean", std_time_s="std")
             .reset_index())
    return agg


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    raw_rows = extract_rows()
    df_raw = pd.DataFrame(raw_rows)

    print(f"\nExtracted {len(df_raw)} timing runs\n")
    print(df_raw.groupby(["date", "test_type"]).size().to_string())

    df_raw.to_parquet(PARQUET_OUT, index=False)
    print(f"\nWrote {PARQUET_OUT.name}")
